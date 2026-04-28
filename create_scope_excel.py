#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script to create Excel file summarizing project scope and progress
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
title_font = Font(bold=True, size=14)
subtitle_font = Font(bold=True, size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center')
wrap_align = Alignment(wrap_text=True, vertical='top')

# ============================================================================
# Sheet 1: Phase Overview
# ============================================================================
ws1 = wb.create_sheet("Phase Overview")
ws1.title = "Phase Overview"

# Header
ws1['A1'] = "Samsung Panich Delivery - Phase Overview"
ws1['A1'].font = Font(bold=True, size=16)
ws1.merge_cells('A1:D1')

# Summary
ws1['A3'] = "ความคืบหน้าโดยรวม: 73% (24/33 งาน)"
ws1['A3'].font = Font(bold=True, size=12)
ws1.merge_cells('A3:D3')

# Phase Status Table
headers = ['Phase', 'ชื่อ Phase', 'ความคืบหน้า', 'สถานะ']
ws1.append(['', '', '', ''])
ws1.append(headers)

# Style header row
for col in range(1, 5):
    cell = ws1.cell(row=5, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# Phase data
phases = [
    ['Phase 1', 'Core Setup', '100% (8/8)', '✅ เสร็จสิ้น'],
    ['Phase 2', 'Basic Features', '100% (10/10)', '✅ เสร็จสิ้น'],
    ['Phase 3', 'Advanced Features', '100% (6/6)', '✅ เสร็จสิ้น'],
    ['Phase 4', 'Security & Compliance', '0% (0/6)', '⏳ ยังไม่เริ่ม'],
    ['Phase 5', 'Testing & Deployment', '0% (0/6)', '⏳ ยังไม่เริ่ม'],
]

for phase in phases:
    ws1.append(phase)

# Style data rows
for row in range(6, 11):
    for col in range(1, 5):
        cell = ws1.cell(row=row, column=col)
        cell.border = border
        if col == 4:  # Status column
            cell.alignment = center_align

# Column widths
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 18
ws1.column_dimensions['D'].width = 18

# ============================================================================
# Sheet 2: Backend Status
# ============================================================================
ws2 = wb.create_sheet("Backend Status")
ws2.title = "Backend Status"

ws2['A1'] = "Backend API Status"
ws2['A1'].font = Font(bold=True, size=16)
ws2.merge_cells('A1:C1')

ws2['A3'] = "Location:"
ws2['B3'] = "/Users/home/Zprojects/shop-delivery"
ws2['A4'] = "API URL:"
ws2['B4'] = "http://127.0.0.1:8000/api/"
ws2['A5'] = "Admin Panel:"
ws2['B5'] = "http://127.0.0.1:8000/admin/"

# Services
ws2.append(['', '', ''])
ws2['A7'] = "API Services"
ws2['A7'].font = subtitle_font
ws2.merge_cells('A7:C7')

headers = ['Service', 'Status', 'Endpoints']
ws2.append(headers)

for col in range(1, 4):
    cell = ws2.cell(row=9, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

services = [
    ['Auth Service', '✅ 100% Complete', '3 endpoints'],
    ['Products Service', '✅ 100% Complete', '2 endpoints'],
    ['Orders Service', '✅ 100% Complete', '8 endpoints'],
    ['Payments Service', '✅ 100% Complete', '6 endpoints'],
    ['Logistics Service', '✅ 100% Complete', '9 endpoints'],
    ['PDPA Service', '✅ 100% Complete', '2 endpoints'],
]

for service in services:
    ws2.append(service)

# Database Models
ws2.append(['', '', ''])
ws2['A16'] = "Database Models"
ws2['A16'].font = subtitle_font
ws2.merge_cells('A16:C16')

headers = ['Model Category', 'Status', 'Details']
ws2.append(headers)

for col in range(1, 4):
    cell = ws2.cell(row=18, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

models = [
    ['User Management', '✅ Complete', 'Customer, Driver, Admin, LINE User'],
    ['Product Management', '✅ Complete', 'Category, Product'],
    ['Order Management', '✅ Complete', 'Order, OrderItem'],
    ['Payment Management', '✅ Complete', 'PaymentTransaction, PaymentWebhook'],
    ['Delivery Management', '✅ Complete', 'Delivery, DriverAssignment'],
    ['PDPA Compliance', '✅ Complete', 'ConsentRecord, AuditLog'],
]

for model in models:
    ws2.append(model)

# Column widths
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 20
ws2.column_dimensions['C'].width = 50

# ============================================================================
# Sheet 3: Frontend Status
# ============================================================================
ws3 = wb.create_sheet("Frontend Status")
ws3.title = "Frontend Status"

ws3['A1'] = "Frontend Status"
ws3['A1'].font = Font(bold=True, size=16)
ws3.merge_cells('A1:C1')

ws3['A3'] = "Location:"
ws3['B3'] = "/Users/home/Zprojects/shop-delivery/frontend"
ws3['A4'] = "Development URL:"
ws3['B4'] = "http://localhost:3000"
ws3['A5'] = "Status:"
ws3['B5'] = "✅ 80% Complete"

# Technology Stack
ws3.append(['', '', ''])
ws3['A7'] = "Technology Stack"
ws3['A7'].font = subtitle_font
ws3.merge_cells('A7:C7')

tech_stack = [
    ['React', '19.2.0'],
    ['React Router DOM', '7.9.4'],
    ['Axios', '1.12.2'],
    ['LIFF SDK', '1.2.0'],
    ['Webpack', '5.102.1'],
]

headers = ['Technology', 'Version']
ws3.append(headers)

for col in range(1, 3):
    cell = ws3.cell(row=9, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

for tech in tech_stack:
    ws3.append(tech)

# Pages
ws3.append(['', '', ''])
ws3['A15'] = "Pages (11 pages)"
ws3['A15'].font = subtitle_font
ws3.merge_cells('A15:C15')

headers = ['Category', 'Pages', 'Status']
ws3.append(headers)

for col in range(1, 4):
    cell = ws3.cell(row=17, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

pages = [
    ['Customer', 'Home, Products, Cart, Checkout, Orders, Tracking, Profile, Login', '✅ Complete'],
    ['Driver', 'DriverLogin', '✅ Complete'],
    ['Admin', 'AdminLogin, AdminDashboard', '✅ Complete'],
]

for page in pages:
    ws3.append(page)

# Components
ws3.append(['', '', ''])
ws3['A21'] = "Components (6 components)"
ws3['A21'].font = subtitle_font
ws3.merge_cells('A21:C21')

ws3['A22'] = "Header, AdminHeader, ProductCard, AddressPicker, LineLoginButton"
ws3['A22'].font = Font(size=10)
ws3.merge_cells('A22:C22')

# Features
ws3.append(['', '', ''])
ws3['A24'] = "Features"
ws3['A24'].font = subtitle_font
ws3.merge_cells('A24:C24')

features = [
    ['API Integration', '✅ Complete (Real API calls via Axios)'],
    ['LINE Login', '✅ Working (LINE OAuth 2.0 Integration)'],
    ['Routing', '✅ Complete (React Router DOM)'],
    ['Styling', '✅ Complete (CSS3, Responsive Design)'],
]

headers = ['Feature', 'Status']
ws3.append(headers)

for col in range(1, 3):
    cell = ws3.cell(row=26, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

for feature in features:
    ws3.append(feature)

# Column widths
ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 50
ws3.column_dimensions['C'].width = 20

# ============================================================================
# Sheet 4: API Endpoints
# ============================================================================
ws4 = wb.create_sheet("API Endpoints")
ws4.title = "API Endpoints"

ws4['A1'] = "API Endpoints Summary"
ws4['A1'].font = Font(bold=True, size=16)
ws4.merge_cells('A1:D1')

headers = ['Service', 'Endpoint', 'Method', 'Description']
ws4.append(['', '', '', ''])
ws4.append(headers)

for col in range(1, 5):
    cell = ws4.cell(row=3, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

endpoints = [
    # Auth Service
    ['Auth Service', '/api/accounts/register/', 'POST', 'ลงทะเบียนลูกค้า'],
    ['Auth Service', '/api/accounts/profile/', 'GET', 'ดูข้อมูลโปรไฟล์'],
    ['Auth Service', '/api/accounts/data-export/', 'GET', 'ส่งออกข้อมูล'],
    # Products Service
    ['Products Service', '/api/products/', 'GET', 'รายการสินค้า'],
    ['Products Service', '/api/products/categories/', 'GET', 'หมวดหมู่สินค้า'],
    # Orders Service
    ['Orders Service', '/api/orders/', 'POST', 'สร้างคำสั่งซื้อ'],
    ['Orders Service', '/api/orders/list/', 'GET', 'รายการคำสั่งซื้อ'],
    ['Orders Service', '/api/orders/cart/add/', 'POST', 'เพิ่มสินค้าในตะกร้า'],
    ['Orders Service', '/api/orders/cart/', 'GET', 'ดูตะกร้าสินค้า'],
    ['Orders Service', '/api/orders/cart/update/', 'PUT', 'แก้ไขสินค้าในตะกร้า'],
    ['Orders Service', '/api/orders/cart/{product_id}/', 'DELETE', 'ลบสินค้าจากตะกร้า'],
    ['Orders Service', '/api/orders/{order_id}/tracking/', 'GET', 'ติดตามคำสั่งซื้อ'],
    ['Orders Service', '/api/orders/{order_id}/status/', 'PUT', 'อัปเดตสถานะคำสั่งซื้อ'],
    # Payments Service
    ['Payments Service', '/api/payments/promptpay/', 'POST', 'สร้าง QR Code PromptPay'],
    ['Payments Service', '/api/payments/truemoney/', 'POST', 'สร้าง QR Code TrueMoney'],
    ['Payments Service', '/api/payments/rabbit/', 'POST', 'สร้าง QR Code Rabbit'],
    ['Payments Service', '/api/payments/scb-easy/', 'POST', 'สร้าง QR Code SCB Easy'],
    ['Payments Service', '/api/payments/status/{transaction_id}/', 'GET', 'เช็คสถานะการชำระเงิน'],
    ['Payments Service', '/api/payments/webhook/', 'POST', 'Payment Webhook'],
    # Logistics Service
    ['Logistics Service', '/api/logistics/calculate-fee/', 'POST', 'คำนวณค่าจัดส่ง'],
    ['Logistics Service', '/api/logistics/driver/assignments/', 'GET', 'งานจัดส่ง'],
    ['Logistics Service', '/api/logistics/driver/assignments/{id}/update/', 'PUT', 'อัปเดตสถานะ'],
    ['Logistics Service', '/api/logistics/drivers/', 'GET', 'รายการคนขับ'],
    ['Logistics Service', '/api/logistics/drivers/', 'POST', 'เพิ่มคนขับ'],
    ['Logistics Service', '/api/logistics/drivers/{id}/', 'GET', 'ดูข้อมูลคนขับ'],
    ['Logistics Service', '/api/logistics/drivers/{id}/', 'PUT', 'แก้ไขข้อมูลคนขับ'],
    ['Logistics Service', '/api/logistics/drivers/{id}/', 'DELETE', 'ลบข้อมูลคนขับ'],
    ['Logistics Service', '/api/logistics/drivers/availability/', 'POST', 'อัปเดตสถานะการพร้อมใช้งาน'],
    # PDPA Service
    ['PDPA Service', '/api/pdpa/privacy-policy/', 'GET', 'นโยบายความเป็นส่วนตัว'],
    ['PDPA Service', '/api/pdpa/consent/', 'GET/POST', 'จัดการความยินยอม'],
]

for endpoint in endpoints:
    ws4.append(endpoint)

# Style data rows
for row in range(4, 34):
    for col in range(1, 5):
        cell = ws4.cell(row=row, column=col)
        cell.border = border
        if col == 3:  # Method column
            cell.alignment = center_align

# Column widths
ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 50
ws4.column_dimensions['C'].width = 12
ws4.column_dimensions['D'].width = 40

# ============================================================================
# Sheet 5: Next Steps
# ============================================================================
ws5 = wb.create_sheet("Next Steps")
ws5.title = "Next Steps"

ws5['A1'] = "Next Steps"
ws5['A1'].font = Font(bold=True, size=16)
ws5.merge_cells('A1:C1')

headers = ['ลำดับ', 'Task', 'Status']
ws5.append(['', '', ''])
ws5.append(headers)

for col in range(1, 4):
    cell = ws5.cell(row=3, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

next_steps = [
    [1, 'เพิ่มรูปสินค้าจริงในฐานข้อมูล', '⏳ Pending'],
    [2, 'Testing - ทดสอบระบบทั้งหมด', '⏳ Pending'],
    [3, 'Deployment - Deploy ไปยัง Production', '⏳ Pending'],
    [4, 'Security & Compliance - Phase 4', '⏳ Pending'],
    [5, 'Testing & Documentation - Phase 5', '⏳ Pending'],
]

for step in next_steps:
    ws5.append(step)

# Style data rows
for row in range(4, 9):
    for col in range(1, 4):
        cell = ws5.cell(row=row, column=col)
        cell.border = border
        if col == 1:  # Number column
            cell.alignment = center_align
        if col == 3:  # Status column
            cell.alignment = center_align

# Column widths
ws5.column_dimensions['A'].width = 10
ws5.column_dimensions['B'].width = 60
ws5.column_dimensions['C'].width = 20

# ============================================================================
# Sheet 6: Project Info
# ============================================================================
ws6 = wb.create_sheet("Project Info")
ws6.title = "Project Info"

ws6['A1'] = "Project Information"
ws6['A1'].font = Font(bold=True, size=16)
ws6.merge_cells('A1:B1')

info = [
    ['Project Name', 'Samsung Panich Delivery System'],
    ['Version', '1.1.0'],
    ['Last Updated', datetime.now().strftime('%Y-%m-%d')],
    ['Backend Path', '/Users/home/Zprojects/shop-delivery'],
    ['Frontend Path', '/Users/home/Zprojects/shop-delivery/frontend'],
    ['Backend URL', 'http://127.0.0.1:8000/api/'],
    ['Frontend URL', 'http://localhost:3000'],
    ['Admin Panel', 'http://127.0.0.1:8000/admin/'],
    ['Progress', '73% (24/33 tasks)'],
    ['Backend Status', '✅ 100% Complete'],
    ['Frontend Status', '✅ 80% Complete'],
]

for i, (key, value) in enumerate(info, start=3):
    ws6.cell(row=i, column=1, value=key).font = Font(bold=True)
    ws6.cell(row=i, column=2, value=value)

ws6.column_dimensions['A'].width = 20
ws6.column_dimensions['B'].width = 60

# Save file
output_file = 'Project_Scope_Summary.xlsx'
wb.save(output_file)
print(f"✅ Created Excel file: {output_file}")


